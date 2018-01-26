#  Copyright 2017 Aaron Zhang.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
import os
from openpyxl import load_workbook


class XlsxExcelLibrary:
    """
    This test library provides keywords to allow opening, reading, writing
     and saving Excel files from Robot Framework.


    *Before running tests*

    Prior to running tests, ExcelLibrary must first be imported into your Robot test suite.

    Example:
        | Library | ExcelLibrary |

    """

    ROBOT_LIBRARY_SCOPE = 'GLOBAL'

    def __init__(self):
        self.wb = None
        self.tb = None
        self.sheetNum = None
        self.sheetNames = None
        self.fileName = None
        if os.name is "nt":
            self.tmpDir = "Temp"
        else:
            self.tmpDir = "tmp"

    def open_excel(self, filename, useTempDir=False):
        """
        Opens the Excel file from the path provided in the file name parameter.
        If the boolean useTempDir is set to true, depending on the operating system of the computer running the test the file will be opened in the Temp directory if the operating system is Windows or tmp directory if it is not.

        Arguments:
                |  File Name (string)                      | The file name string value that will be used to open the excel file to perform tests upon.                                  |
                |  Use Temporary Directory (default=False) | The file will not open in a temporary directory by default. To activate and open the file in a temporary directory, pass 'True' in the variable. |
        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |

        """
        if useTempDir is True:
            print('Opening file at %s' % filename)
            self.wb = load_workbook(os.path.join("/", self.tmpDir, filename))
        else:
            self.wb = load_workbook(filename)
        self.fileName = filename
        self.sheetNames = self.wb.get_sheet_names()
        # print self.wb.get_sheet_names()

    def open_excel_current_directory(self, filename):
        """
        Opens the Excel file from the current directory using the directory the test has been run from.

        Arguments:
                |  File Name (string)  | The file name string value that will be used to open the excel file to perform tests upon.  |
        Example:

        | *Keywords*           |  *Parameters*        |
        | Open Excel           |  ExcelRobotTest.xls  |

        """
        workdir = os.getcwd()
        print('Opening file at %s' % filename)
        self.wb = load_workbook(os.path.join(workdir, filename))
        self.sheetNames = self.wb.get_sheet_names()

    def get_sheet_names(self):
        """
        Returns the names of all the worksheets in the current workbook.

        Example:

        | *Keywords*              |  *Parameters*                                      |
        | Open Excel              |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Get Sheets Names        |                                                    |

        """
        sheetnames = self.wb.get_sheet_names()
        return sheetnames

    def get_number_of_sheets(self):
        """
        Returns the number of worksheets in the current workbook.

        Example:

        | *Keywords*              |  *Parameters*                                      |
        | Open Excel              |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Get Number of Sheets    |                                                    |

        """
        sheetnum = len(self.wb.get_sheet_names())
        return sheetnum

    def get_column_count(self, sheetname):
        """
        Returns the specific number of columns of the sheet name specified.

        Arguments:
                |  Sheet Name (string)  | The selected sheet that the column count will be returned from. |
        Example:

        | *Keywords*          |  *Parameters*                                      |
        | Open Excel          |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Get Column Count    |  TestSheet1                                        |

        """
        xlxswb = self.wb[sheetname]
        return len(tuple(xlxswb.columns))

    def get_row_count(self, sheetname):
        """
        Returns the specific number of rows of the sheet name specified.

        Arguments:
                |  Sheet Name (string)  | The selected sheet that the row count will be returned from. |
        Example:

        | *Keywords*          |  *Parameters*                                      |
        | Open Excel          |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Get Row Count       |  TestSheet1                                        |

        """
        xlxswb = self.wb[sheetname]
        return len(tuple(xlxswb.rows))

    def get_column_values(self, sheetname, columnname, includeEmptyCells=True):
        """
        Returns the specific column values of the sheet name specified.

        Arguments:
                |  Sheet Name (string)                 | The selected sheet that the column values will be returned from.                                                            |
                |  Column (int)                        | The column integer value that will be used to select the column from which the values will be returned.                     |
                |  Include Empty Cells (default=True)  | The empty cells will be included by default. To deactivate and only return cells with values, pass 'False' in the variable. |
        Example:

        | *Keywords*           |  *Parameters*                                          |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |   |
        | Get Column Values    |  TestSheet1                                        | 0 |

        """
        vals = []
        xlxswb = self.wb[sheetname]
        for cell in tuple(xlxswb[columnname]):
            vals.append(cell.value)
        return vals

    def get_row_values(self, sheetname, row, includeEmptyCells=True):
        """
        Returns the specific row values of the sheet name specified.

        Arguments:
                |  Sheet Name (string)                 | The selected sheet that the row values will be returned from.                                                               |
                |  Row (int)                           | The row integer value that will be used to select the row from which the values will be returned.                           |
                |  Include Empty Cells (default=True)  | The empty cells will be included by default. To deactivate and only return cells with values, pass 'False' in the variable. |
        Example:

        | *Keywords*           |  *Parameters*                                          |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |   |
        | Get Row Values       |  TestSheet1                                        | 0 |

        """
        vals = []
        xlxswb = self.wb[sheetname]
        for cell in tuple(xlxswb[int(row)]):
            vals.append(cell.value)
        return vals

    # TODO: change to return correct values
    def get_sheet_values(self, sheetname, includeEmptyCells=True):
        """
        Returns the values from the sheet name specified.

        Arguments:
                |  Sheet Name (string)                 | The selected sheet that the cell values will be returned from.                                                              |
                |  Include Empty Cells (default=True)  | The empty cells will be included by default. To deactivate and only return cells with values, pass 'False' in the variable. |
        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Get Sheet Values     |  TestSheet1                                        |

        """
        xlxswb = self.wb[sheetname]
        return xlxswb

    # TODO: need to change the correct return value
    def get_workbook_values(self, includeEmptyCells=True):
        """
        Returns the values from each sheet of the current workbook.

        Arguments:
                |  Include Empty Cells (default=True)  | The empty cells will be included by default. To deactivate and only return cells with values, pass 'False' in the variable. |
        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Get Workbook Values  |                                                    |

        """

    def read_cell_data_by_name(self, sheetname, cell_name):
        """
        Uses the cell name to return the data from that cell.

        Arguments:
                |  Sheet Name (string)  | The selected sheet that the cell value will be returned from.  |
                |  Cell Name (string)   | The selected cell name that the value will be returned from.   |
        Example:

        | *Keywords*           |  *Parameters*                                             |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |      |
        | Get Cell Data        |  TestSheet1                                        |  A2  |

        """
        return self.wb[sheetname][cell_name].value

    def read_cell_data_by_coordinates(self, sheetname, column, row):
        """
        Uses the column and row to return the data from that cell.

        Arguments:
                |  Sheet Name (string)  | The selected sheet that the cell value will be returned from.         |
                |  Column (int)         | The column integer value that the cell value will be returned from.   |
                |  Row (int)            | The row integer value that the cell value will be returned from.      |
        Example:

        | *Keywords*     |  *Parameters*                                              |
        | Open Excel     |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |   |   |
        | Read Cell      |  TestSheet1                                        | 0 | 0 |

        """
        cellvalue = self.wb[sheetname].cell(coordinate=None, row=int(row), column=int(column), value=None).value
        return cellvalue

    # TODO: need to change for check file type
    def check_cell_type(self, sheetname, column, row):
        """
        Checks the type of value that is within the cell of the sheet name selected.

        Arguments:
                |  Sheet Name (string)  | The selected sheet that the cell type will be checked from.          |
                |  Column (int)         | The column integer value that will be used to check the cell type.   |
                |  Row (int)            | The row integer value that will be used to check the cell type.      |
        Example:

        | *Keywords*           |  *Parameters*                                              |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |   |   |
        | Check Cell Type      |  TestSheet1                                        | 0 | 0 |

        """

    def put_value_to_cell(self, sheetname, column, row, value):
        """
        Using the sheet name the value of the indicated cell is set to be the number given in the parameter.

        Arguments:
                |  Sheet Name (string) | The selected sheet that the cell will be modified from.                                           |
                |  Column (int)        | The column integer value that will be used to modify the cell.                                    |
                |  Row (int)           | The row integer value that will be used to modify the cell.                                       |
                |  Value (int)         | The integer value that will be added to the specified sheetname at the specified column and row.  |
        Example:

        | *Keywords*           |  *Parameters*                                                         |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |     |     |      |
        | Put Number To Cell   |  TestSheet1                                        |  0  |  0  |  34  |

        """
        putcell = self.wb[sheetname].cell(row=int(row), column=int(column))
        putcell.value = value

    def put_number_to_cell(self, sheetname, column, row, value):
        """
        Using the sheet name the value of the indicated cell is set to be the number given in the parameter.

        Arguments:
                |  Sheet Name (string) | The selected sheet that the cell will be modified from.                                           |
                |  Column (int)        | The column integer value that will be used to modify the cell.                                    |
                |  Row (int)           | The row integer value that will be used to modify the cell.                                       |
                |  Value (int)         | The integer value that will be added to the specified sheetname at the specified column and row.  |
        Example:

        | *Keywords*           |  *Parameters*                                                         |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |     |     |      |
        | Put Number To Cell   |  TestSheet1                                        |  0  |  0  |  34  |

        """
        self.put_value_to_cell(sheetname, column, row, value)

    def put_string_to_cell(self, sheetname, column, row, value):
        """
        Using the sheet name the value of the indicated cell is set to be the string given in the parameter.

        Arguments:
                |  Sheet Name (string) | The selected sheet that the cell will be modified from.                                           |
                |  Column (int)        | The column integer value that will be used to modify the cell.                                    |
                |  Row (int)           | The row integer value that will be used to modify the cell.                                       |
                |  Value (string)      | The string value that will be added to the specified sheetname at the specified column and row.   |
        Example:

        | *Keywords*           |  *Parameters*                                                           |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |     |     |        |
        | Put String To Cell   |  TestSheet1                                        |  0  |  0  |  Hello |

        """
        self.put_value_to_cell(sheetname, column, row, value)

    def save_excel(self, filename, useTempDir=False):
        """
        Saves the Excel file indicated by file name, the useTempDir can be set to true if the user needs the file saved in the temporary directory.
        If the boolean useTempDir is set to true, depending on the operating system of the computer running the test the file will be saved in the Temp directory if the operating system is Windows or tmp directory if it is not.

        Arguments:
                |  File Name (string)                      | The name of the of the file to be saved.  |
                |  Use Temporary Directory (default=False) | The file will not be saved in a temporary directory by default. To activate and save the file in a temporary directory, pass 'True' in the variable. |
        Example:

        | *Keywords*           |  *Parameters*                                      |
        | Open Excel           |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Save Excel           |  NewExcelRobotTest.xls                             |

        """
        if useTempDir is True:
            print('*DEBUG* Got fname %s' % filename)
            self.wb.save(os.path.join("/", self.tmpDir, filename))
        else:
            self.wb.save(filename)

    def save(self):
        """
        Saves the Excel file if the workbook already opened
        """
        self.wb.save(self.fileName)

    def save_excel_current_directory(self, filename):
        """
        Saves the Excel file from the current directory using the directory the test has been run from.

        Arguments:
                |  File Name (string)    | The name of the of the file to be saved.  |
        Example:

        | *Keywords*                     |  *Parameters*                                      |
        | Open Excel                     |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xls  |
        | Save Excel Current Directory   |  NewTestCases.xls                                  |

        """
        workdir = os.getcwd()
        print('*DEBUG* Got fname %s' % filename)
        self.wb.save(os.path.join(workdir, filename))
